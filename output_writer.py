"""Excel output writer matching the SQL database import schema.

Generates .xlsx files with:
- Exact column layout matching the import process
- Colour-coded highlighting for confidence levels
- Bloomberg-style ticker format (XXX LN)
- Dates as datetime objects formatted DD-MMM-YYYY
- NOT_TRACKED tickers on a separate "Not Tracked" sheet
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

import config
from extractor import Extractor


# Highlight colours (ARGB hex)
FILL_YELLOW = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')      # Claude filled
FILL_ORANGE = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')      # Disagreement
FILL_RED = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')          # Math mismatch / both failed
FILL_LIGHT_BLUE = PatternFill(start_color='FFADD8E6', end_color='FFADD8E6', fill_type='solid')  # Always-review verified
FILL_LIGHT_RED = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')    # Missing required field
FILL_GREY = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')         # Not tracked
FILL_LAVENDER = PatternFill(start_color='FFE6CCFF', end_color='FFE6CCFF', fill_type='solid')   # Flagged — needs manual review

# Column layout matching SQL import
COLUMNS = [
    'Ticker',            # A
    'Event Type',        # B
    'Price',             # C
    'Cancelled shares',  # D (always blank)
    'Treasury shares',   # E (buyback shares)
    'Issued Shares',     # F (issuance shares)
    'Announce date',     # G
    'Transaction date',  # H
    'Description',       # I
    'New shares in issue',  # J
    'Announcement',      # K
]


class OutputWriter:
    """Write extraction results to Excel in SQL import format."""

    @staticmethod
    def write(results, output_path, highlights=None, validations=None):
        """Write results to Excel file.

        Args:
            results: list of dicts from extraction
            output_path: path for the .xlsx file
            highlights: dict {ticker: {claude_filled: [...], disagreement: [...]}}
            validations: dict {ticker: {valid: bool, ...}} from reconciler
        """
        wb = openpyxl.Workbook()

        # Split results into tracked and not-tracked
        not_tracked_set = set(getattr(config, 'NOT_TRACKED_TICKERS', []))
        tracked_results = [r for r in results if r.get('ticker', '') not in not_tracked_set]
        not_tracked_results = [r for r in results if r.get('ticker', '') in not_tracked_set]

        # Write main "output" sheet with tracked results
        ws = wb.active
        ws.title = config.OUTPUT_SHEET_NAME
        _write_sheet(ws, tracked_results, highlights, validations)

        # Write "Not Tracked" sheet if there are untracked tickers
        if not_tracked_results:
            ws_nt = wb.create_sheet(title='Not Tracked')
            _write_sheet(ws_nt, not_tracked_results, highlights, validations,
                         row_fill=FILL_GREY)

        wb.save(output_path)
        print(f"Saved output to {output_path}")
        if not_tracked_results:
            print(f"  ({len(tracked_results)} tracked, "
                  f"{len(not_tracked_results)} on 'Not Tracked' sheet)")
        return output_path


def _write_sheet(ws, results, highlights, validations, row_fill=None):
    """Write results to a worksheet with headers, data, and highlighting.

    Args:
        ws: openpyxl Worksheet
        results: list of result dicts
        highlights: dict {ticker: {claude_filled: [...], ...}}
        validations: dict {ticker_ln: {valid: bool, ...}}
        row_fill: optional PatternFill to apply to entire row (e.g. grey for not-tracked)
    """
    # Write header row
    header_font = Font(bold=True)
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    # Write data rows
    for row_idx, data in enumerate(results, 2):
        ticker = data.get('ticker', '')
        # Add "LN" suffix unless ticker already has an exchange suffix (e.g. "VCVOF US")
        ticker_ln = ticker if ' ' in ticker else f"{ticker} LN"
        tx_type = data.get('transaction_type', '')

        # Determine shares column placement
        shares = data.get('shares_transacted')
        treasury_shares = shares if tx_type == 'Buyback' and shares else None
        issued_shares = shares if tx_type == 'Issuance' and shares else None

        # Format dates as datetime objects for proper Excel date handling
        announce_date = Extractor.parse_date_to_datetime(data.get('announcement_date'))
        effective_date = Extractor.parse_date_to_datetime(data.get('effective_date'))

        # Use voting_rights as shares in issue (preferred), fall back to shares_outstanding
        shares_in_issue = data.get('voting_rights') or data.get('shares_outstanding')

        # Page text for Announcement column
        page_text = data.get('page_text', '')

        # Write row
        row_data = [
            ticker_ln,          # A: Ticker
            tx_type,            # B: Event Type
            data.get('average_price'),  # C: Price
            None,               # D: Cancelled shares (always blank)
            treasury_shares,    # E: Treasury shares
            issued_shares,      # F: Issued Shares
            announce_date,      # G: Announce date
            effective_date,     # H: Transaction date
            data.get('extraction_method') if data.get('extraction_method', '').startswith(('FAILED', 'FLAGGED')) else config.OUTPUT_DESCRIPTION,  # I: Description
            shares_in_issue,    # J: New shares in issue
            page_text,          # K: Announcement
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply DD-MMM-YYYY format to date columns
        for date_col in (7, 8):  # G=Announce date, H=Transaction date
            cell = ws.cell(row=row_idx, column=date_col)
            if cell.value is not None:
                cell.number_format = 'DD-MMM-YYYY'

        # Apply row-level fill first (e.g. grey for not-tracked)
        if row_fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Duplicate-flagged rows: lavender fill + note in Description
        if data.get('duplicate_flag'):
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_LAVENDER
            desc_cell = ws.cell(row=row_idx, column=9)  # I: Description
            existing = desc_cell.value or ''
            desc_cell.value = f"REVIEW: Multiple announcements for this ticker on same day. {existing}".strip()

        # Flagged rows: lavender fill for manual-review items
        extraction_method = data.get('extraction_method', '') or ''
        if extraction_method.startswith('FLAGGED'):
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_LAVENDER
            continue  # Skip normal highlighting for flagged rows

        # Apply highlighting (only if no row_fill override)
        if not row_fill:
            ticker_highlights = (highlights or {}).get(ticker, {})
            ticker_validation = (validations or {}).get(ticker_ln, {})

            # Red: math mismatch
            if ticker_validation and not ticker_validation.get('skipped') and not ticker_validation.get('valid'):
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = FILL_RED

            # Orange: disagreement (on specific cells)
            for field in ticker_highlights.get('disagreement', []):
                col = _field_to_column(field, tx_type)
                if col:
                    ws.cell(row=row_idx, column=col).fill = FILL_ORANGE

            # Yellow: Claude filled (on specific cells)
            for field in ticker_highlights.get('claude_filled', []):
                col = _field_to_column(field, tx_type)
                if col:
                    ws.cell(row=row_idx, column=col).fill = FILL_YELLOW

            # Light blue: always-review verified (whole row)
            if ticker in config.ALWAYS_REVIEW_TICKERS and not ticker_highlights.get('disagreement'):
                for col_idx in range(1, len(COLUMNS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill == PatternFill():  # only if not already highlighted
                        cell.fill = FILL_LIGHT_BLUE

            # Light red: missing required fields (Price, shares column, New shares in issue)
            # Overrides light blue (informational) but not red/orange/yellow (data-quality)
            _OVERRIDABLE_FILLS = {'00000000', 'FFADD8E6'}  # default, light blue
            shares_col = 6 if tx_type == 'Issuance' else 5  # E=Treasury, F=Issued
            for req_col in (3, shares_col, 10):  # C=Price, E/F=shares, J=shares in issue
                cell = ws.cell(row=row_idx, column=req_col)
                fill_rgb = getattr(cell.fill.fgColor, 'rgb', '00000000') or '00000000'
                if cell.value is None and fill_rgb in _OVERRIDABLE_FILLS:
                    cell.fill = FILL_LIGHT_RED

    # Auto-adjust column widths (except Announcement which can be very long)
    for col_idx in range(1, len(COLUMNS)):  # skip last column (Announcement)
        max_width = len(COLUMNS[col_idx - 1])
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_width = max(max_width, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_width + 2, 30)

    # Announcement column gets a fixed width
    ws.column_dimensions['K'].width = 50


def _field_to_column(field_name, transaction_type=None):
    """Map extraction field name to Excel column number."""
    mapping = {
        'average_price': 3,         # C: Price
        'announcement_date': 7,     # G: Announce date
        'effective_date': 8,        # H: Transaction date
        'transaction_type': 2,      # B: Event Type
        'voting_rights': 10,        # J: New shares in issue
        'shares_outstanding': 10,   # J: New shares in issue
    }
    # Shares column depends on transaction type
    if field_name == 'shares_transacted':
        return 6 if transaction_type == 'Issuance' else 5
    return mapping.get(field_name)
