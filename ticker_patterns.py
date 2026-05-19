"""Ticker-specific extraction patterns loaded from Excel database.

The Excel file contains start/end text markers for each ticker's announcement
format. These markers allow precise extraction of fields like price, shares,
voting rights by finding text between known delimiters.

Falls back gracefully if the Excel file is missing — the scraper will use
generic regex extraction instead.
"""

import pandas as pd
import re


class TickerPatternsDB:
    """Load and apply ticker-specific extraction patterns from Excel."""

    def __init__(self, excel_path=None):
        self.patterns = {}
        self.loaded = False
        self._excel_path = excel_path
        if excel_path:
            self.load_from_excel(excel_path)

    def load_from_excel(self, excel_path):
        """Load patterns from Excel file with 'db' sheet and multi-level headers."""
        try:
            df = pd.read_excel(excel_path, sheet_name='db', header=[0, 1])
            for _, row in df.iterrows():
                ticker = str(row.iloc[0]).strip()
                if not ticker or ticker == 'nan':
                    continue
                ticker_clean = ticker.replace(' LN', '').strip()
                if not ticker_clean:
                    continue

                self.patterns[ticker_clean] = {
                    'shares_start': str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else '',
                    'shares_end': str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else '',
                    'price_start': str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else '',
                    'price_end': str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else '',
                    'total_shares_start': str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else '',
                    'total_shares_end': str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else '',
                    'treasury_start': str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else '',
                    'treasury_end': str(row.iloc[10]).strip() if pd.notna(row.iloc[10]) else '',
                    'voting_rights_start': str(row.iloc[11]).strip() if pd.notna(row.iloc[11]) else '',
                    'voting_rights_end': str(row.iloc[12]).strip() if pd.notna(row.iloc[12]) else '',
                    'currency': str(row.iloc[13]).strip() if len(row) > 13 and pd.notna(row.iloc[13]) else 'GBp',
                }
            self.loaded = True
            print(f"  ✓ Loaded patterns for {len(self.patterns)} tickers")
        except Exception as e:
            print(f"  ⚠ Could not load ticker patterns: {e}")
            self.loaded = False

    def has_patterns(self, ticker):
        ticker_clean = ticker.replace(' LN', '').strip()
        return ticker_clean in self.patterns

    def get_patterns(self, ticker):
        ticker_clean = ticker.replace(' LN', '').strip()
        return self.patterns.get(ticker_clean)

    def extract_between(self, text, start_marker, end_marker):
        """Extract text substring between two markers.

        Uses the last occurrence of both markers so that when a marker word
        appears in a preamble sentence *and* as a table label, the table
        label (closest to the value) wins.
        """
        if not start_marker or start_marker == 'nan' or not end_marker or end_marker == 'nan':
            return None
        start_marker_lower = start_marker.lower()
        end_marker_lower = end_marker.lower()
        text_lower = text.lower()

        # Use last occurrence of end marker, then last start marker before it
        end_idx = text_lower.rfind(end_marker_lower)
        if end_idx == -1:
            return None

        start_idx = text_lower.rfind(start_marker_lower, 0, end_idx)
        if start_idx == -1:
            return None
        start_idx += len(start_marker_lower)

        return text[start_idx:end_idx].strip()

    def parse_number(self, text_with_number):
        """Extract first numeric value from text, handling commas."""
        if not text_with_number:
            return None
        cleaned = text_with_number.replace(',', '')
        match = re.search(r'[\d]+\.?[\d]*', cleaned)
        if match:
            num_str = match.group()
            if '.' in num_str:
                return float(num_str)
            return int(num_str)
        return None

    def extract_with_patterns(self, ticker, text):
        """Extract all fields using ticker-specific start/end markers.

        Returns dict with extracted fields, or None if extraction fails.
        The text is stripped of whitespace before marker matching.
        """
        patterns = self.get_patterns(ticker)
        if not patterns:
            return None

        # Remove all whitespace for marker matching (critical for DB patterns)
        text_no_ws = re.sub(r'\s+', '', text)

        result = {
            'shares_transacted': None,
            'average_price': None,
            'voting_rights': None,
            'shares_outstanding': None,
            'held_in_treasury': None,
            'currency': patterns.get('currency', 'GBp'),
        }

        field_map = {
            'shares_transacted': ('shares_start', 'shares_end'),
            'average_price': ('price_start', 'price_end'),
            'voting_rights': ('voting_rights_start', 'voting_rights_end'),
            'shares_outstanding': ('total_shares_start', 'total_shares_end'),
            'held_in_treasury': ('treasury_start', 'treasury_end'),
        }

        for field, (start_key, end_key) in field_map.items():
            extracted = self.extract_between(text_no_ws, patterns[start_key], patterns[end_key])
            if extracted:
                num = self.parse_number(extracted)
                if num is not None:
                    result[field] = num

        if result['shares_transacted'] is None and result['average_price'] is None:
            return None

        return result

    def write_pattern(self, ticker, patterns_dict):
        """Write or update a ticker's patterns in the Excel database.

        Uses openpyxl (already a dependency via pandas) to write directly.
        Column indices are 1-based (openpyxl convention), matching the 0-based
        iloc offsets in load_from_excel + 1. Data rows start at row 3 because
        pd.read_excel header=[0, 1] consumes the first 2 rows as multi-level headers.
        """
        if not self.loaded or not self._excel_path:
            print(f"    ⚠ Cannot write pattern for {ticker} — patterns DB not loaded")
            return False

        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._excel_path, keep_vba=True)
            ws = wb['db']

            target_row = None
            first_empty = None
            for row_idx in range(3, ws.max_row + 2):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val:
                    clean = str(cell_val).replace(' LN', '').strip()
                    if clean == ticker:
                        target_row = row_idx
                        break
                elif first_empty is None:
                    first_empty = row_idx

            if target_row is None:
                target_row = first_empty or ws.max_row + 1
                ws.cell(row=target_row, column=1, value=f"{ticker} LN")

            col_map = {
                'shares_start': 4, 'shares_end': 5,
                'price_start': 6, 'price_end': 7,
                'total_shares_start': 8, 'total_shares_end': 9,
                'treasury_start': 10, 'treasury_end': 11,
                'voting_rights_start': 12, 'voting_rights_end': 13,
                'currency': 14,
            }

            for key, col in col_map.items():
                val = patterns_dict.get(key, '')
                if val:
                    ws.cell(row=target_row, column=col, value=val)

            wb.save(self._excel_path)
            wb.close()

            existing = self.patterns.get(ticker, {})
            existing.update(patterns_dict)
            self.patterns[ticker] = existing
            print(f"    ★ Auto-learned patterns for {ticker} written to DB")
            return True

        except Exception as e:
            print(f"    ⚠ Failed to write pattern for {ticker}: {e}")
            return False
